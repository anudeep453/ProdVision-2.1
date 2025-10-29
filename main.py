import os
import threading
import time
import random
import logging

from flask import Flask, render_template, request, jsonify, session, send_file
from flask_session import Session
from datetime import datetime, timedelta
import bcrypt
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
# Import new independent row adapter
from independent_row_adapter import EntryManager as ProductionEntryManager
from config import SECRET_KEY, DEBUG, HOST, PORT

# Centralized logging configuration
# We configure logging early so any subsequent module imports can use it.
LOG_LEVEL_NAME = os.getenv("PRODVISION_LOG_LEVEL")  # Optional override (e.g., INFO, DEBUG, WARNING)

def _resolve_log_level():
    if LOG_LEVEL_NAME:
        return getattr(logging, LOG_LEVEL_NAME.upper(), logging.INFO)
    # Fall back to DEBUG flag: DEBUG -> INFO (not too noisy) else WARNING
    return logging.INFO if DEBUG else logging.WARNING

logging.basicConfig(
    level=_resolve_log_level(),
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

logger = logging.getLogger("prodvision")

# Reduce verbosity of Werkzeug request logs unless explicit DEBUG requested
if not DEBUG and not LOG_LEVEL_NAME:
    logging.getLogger('werkzeug').setLevel(logging.ERROR)

app = Flask(__name__)

# Enable CORS for credentials
CORS(app, supports_credentials=True)

# Configuration
app.config['SECRET_KEY'] = SECRET_KEY
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = './flask_session'  # Explicitly set session directory
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'prodvision:'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

# Initialize extensions
Session(app)

# Initialize SharePoint SQLite database manager
entry_manager = ProductionEntryManager()

# Global lock to guard against race conditions creating duplicate date/application entries
_create_entry_lock = threading.Lock()

# Session cleanup functions
def cleanup_expired_session_files():
    """Clean up expired and orphaned session files"""
    try:
        session_dir = app.config['SESSION_FILE_DIR']
        if not os.path.exists(session_dir):
            return 0
        
        current_time = time.time()
        session_lifetime_seconds = app.config['PERMANENT_SESSION_LIFETIME'].total_seconds()
        deleted_count = 0
        
        for filename in os.listdir(session_dir):
            file_path = os.path.join(session_dir, filename)
            
            # Skip directories
            if not os.path.isfile(file_path):
                continue
            
            # Get file modification time
            file_mtime = os.path.getmtime(file_path)
            file_age = current_time - file_mtime
            
            # Delete if file is older than session lifetime
            if file_age > session_lifetime_seconds:
                try:
                    os.remove(file_path)
                    deleted_count += 1
                    logger.debug("Deleted expired session file: %s", filename)
                except OSError as e:
                    logger.warning("Error deleting session file %s: %s", filename, e)
        
        if deleted_count > 0:
            logger.info("Session cleanup completed: %d files deleted", deleted_count)
        
        return deleted_count
    except Exception as e:
        logger.error("Session cleanup error: %s", e, exc_info=True)
        return 0

def delete_current_session_file():
    """Delete the current user's session file immediately"""
    try:
        # Get the current session ID
        session_id = session.get('_id')
        if not session_id:
            # Try to get session ID from Flask-Session's internal storage
            session_id = getattr(session, '_id', None)
        
        if session_id:
            # Remove the session key prefix if present
            session_key = app.config['SESSION_KEY_PREFIX'] + session_id
            session_file_path = os.path.join(app.config['SESSION_FILE_DIR'], session_key)
            
            if os.path.exists(session_file_path):
                os.remove(session_file_path)
                logger.debug("Immediately deleted session file: %s", session_key)
                return True
        return False
    except Exception as e:
        logger.error("Error deleting current session file: %s", e, exc_info=True)
        return False

def check_and_cleanup_expired_sessions():
    """Check for expired sessions and clean them immediately"""
    try:
        session_dir = app.config['SESSION_FILE_DIR']
        if not os.path.exists(session_dir):
            return 0
        
        current_time = time.time()
        session_lifetime_seconds = app.config['PERMANENT_SESSION_LIFETIME'].total_seconds()
        deleted_count = 0
        
        for filename in os.listdir(session_dir):
            file_path = os.path.join(session_dir, filename)
            
            if not os.path.isfile(file_path):
                continue
            
            # Check if file is expired
            file_mtime = os.path.getmtime(file_path)
            file_age = current_time - file_mtime
            
            if file_age > session_lifetime_seconds:
                try:
                    os.remove(file_path)
                    deleted_count += 1
                    logger.debug("Immediately deleted expired session file: %s", filename)
                except OSError as e:
                    logger.warning("Error deleting expired session file %s: %s", filename, e)
        
        return deleted_count
    except Exception as e:
        logger.error("Error in immediate session cleanup: %s", e, exc_info=True)
        return 0

def periodic_session_cleanup():
    """Run session cleanup every 30 minutes"""
    while True:
        time.sleep(1800)  # 30 minutes
        cleanup_expired_session_files()

def get_session_stats():
    """Get session file statistics"""
    try:
        session_dir = app.config['SESSION_FILE_DIR']
        if not os.path.exists(session_dir):
            return {'total_files': 0, 'total_size': 0}
        
        total_files = 0
        total_size = 0
        
        for filename in os.listdir(session_dir):
            file_path = os.path.join(session_dir, filename)
            if os.path.isfile(file_path):
                total_files += 1
                total_size += os.path.getsize(file_path)
        
        return {
            'total_files': total_files,
            'total_size': total_size,
            'total_size_mb': round(total_size / (1024 * 1024), 2)
        }
    except Exception as e:
        logger.error("Error getting session stats: %s", e, exc_info=True)
        return {'total_files': 0, 'total_size': 0}

# Helper functions for data validation and conversion
def validate_entry_data(data):
    """Validate production entry data based on application type.
    NOTE: OTHERS should behave like XVA/REG for required fields (only date & application_name).
    """
    application_name = data.get('application_name', '')

    # Common required fields for all applications
    common_required_fields = ['date', 'application_name']

    # Define required fields based on application type
    if application_name in ('XVA', 'REG', 'OTHERS'):
        # Minimal requirement set
        required_fields = common_required_fields
    else:
        # CVAR (ALL/NYQ) fields: prc_mail_text and prc_mail_status are now optional
        required_fields = common_required_fields
    
    # Check required fields (allow arrays for multiple items)
    for field in required_fields:
        if field not in data or not data[field]:
            return False, f'Missing required field: {field}'
    
    # Validate status values based on application type
    if application_name == 'XVA':
        # XVA-specific validation
        if data.get('valo_status') and data['valo_status'] not in ['Red', 'Yellow', 'Green']:
            return False, 'Invalid VALO status'
        if data.get('sensi_status') and data['sensi_status'] not in ['Red', 'Yellow', 'Green']:
            return False, 'Invalid SENSI status'
        if data.get('cf_ra_status') and data['cf_ra_status'] not in ['Red', 'Yellow', 'Green']:
            return False, 'Invalid CF RA status'
        if data.get('quality_legacy') and data['quality_legacy'] not in ['Red', 'Yellow', 'Green']:
            return False, 'Invalid quality legacy status'
        if data.get('quality_target') and data['quality_target'] not in ['Red', 'Yellow', 'Green']:
            return False, 'Invalid quality target status'
        # Skip CVAR-specific validation for XVA entries
    elif application_name == 'REG':
        # REG-specific validation (accept legacy and new values)
        reg_status_val = data.get('reg_status')
        if reg_status_val:
            valid_reg_statuses = {'ongoing', 'open', 'closed', 'Open', 'In Progress', 'Resolved', 'Closed'}
            if reg_status_val not in valid_reg_statuses:
                return False, 'Invalid REG status'
        # Skip CVAR/XVA-specific validation for REG entries
    elif application_name == 'OTHERS':
        # OTHERS-specific validation - only date and application_name are required
    # All other fields (BUSINESS CHAIN, TIMINGS, PUNTUALITY_ISSUE, QUALITY, QUALITY_ISSUE, PRB, HIIM) are optional
        pass
    else:
        # CVAR-specific validation - validate single fields or arrays
        if data.get('prc_mail_status') and data['prc_mail_status'] not in ['Red', 'Yellow', 'Green']:
            return False, 'Invalid PRC mail status'
        if data.get('cp_alerts_status') and data['cp_alerts_status'] not in ['Red', 'Yellow', 'Green']:
            return False, 'Invalid CP alerts status'
        if data.get('quality_status') and data['quality_status'] not in ['Red', 'Yellow', 'Green']:
            return False, 'Invalid quality status'

        # Validate PRBs array if present
        if 'prbs' in data and isinstance(data['prbs'], list):
            for prb in data['prbs']:
                if prb is not None and 'prb_id_number' in prb and prb['prb_id_number'] is not None:
                    try:
                        int(prb['prb_id_number'])
                    except Exception:
                        return False, 'Invalid PRB id number'
                if prb is not None and prb.get('prb_id_status') and prb['prb_id_status'] not in ['active', 'closed']:
                    return False, 'Invalid PRB ID status in array'

        # Validate HIIMs array if present
        if 'hiims' in data and isinstance(data['hiims'], list):
            for hiim in data['hiims']:
                if hiim is not None and 'hiim_id_number' in hiim and hiim['hiim_id_number'] is not None:
                    try:
                        int(hiim['hiim_id_number'])
                    except Exception:
                        return False, 'Invalid HIIM id number'
                if hiim is not None and hiim.get('hiim_id_status') and hiim['hiim_id_status'] not in ['active', 'closed']:
                    return False, 'Invalid HIIM ID status in array'
    
    # Common validation for all applications
    if data.get('prb_id_status') and data['prb_id_status'] not in ['active', 'closed']:
        return False, 'Invalid PRB ID status'
    if data.get('hiim_id_status') and data['hiim_id_status'] not in ['active', 'closed']:
        return False, 'Invalid HIIM ID status'
    
    return True, None

def convert_date_string(date_str):
    """Convert date string to datetime object"""
    if isinstance(date_str, str):
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    return date_str

def validate_independent_row_constraints(data):
    """
    Validate that independent row data doesn't accidentally couple rows across dates
    Each row must be completely independent with its own date and application
    """
    # Ensure date is present and valid
    if not data.get('date'):
        return False, 'Date is required for independent row validation'
    
    # Ensure application_name is present
    if not data.get('application_name'):
        return False, 'Application name is required for independent row validation'
    
    # Validate date format
    try:
        convert_date_string(data['date'])
    except ValueError:
        return False, 'Invalid date format. Use YYYY-MM-DD'
    
    # Check if attempting to create multiple entries with different dates in one call
    # This shouldn't happen in normal operation but validates against API misuse
    if isinstance(data.get('prbs'), list):
        for prb in data['prbs']:
            if prb is not None and prb.get('date') and prb['date'] != data['date']:
                return False, 'All PRBs must have the same date as the main entry for independent row integrity'
    
    if isinstance(data.get('hiims'), list):
        for hiim in data['hiims']:
            if hiim is not None and hiim.get('date') and hiim['date'] != data['date']:
                return False, 'All HIIMs must have the same date as the main entry for independent row integrity'
    
    if isinstance(data.get('issues'), list):
        for issue in data['issues']:
            if issue is not None and issue.get('date') and issue['date'] != data['date']:
                return False, 'All Issues must have the same date as the main entry for independent row integrity'
    
    # Validate that row_type if provided is valid
    if data.get('row_type') and data['row_type'] not in ['main', 'prb', 'hiim', 'issue']:
        return False, 'Invalid row_type. Must be one of: main, prb, hiim, issue'
    
    return True, None

# Authentication helper functions
def is_authenticated():
    # Check for expired sessions and clean them immediately
    check_and_cleanup_expired_sessions()
    
    return 'authenticated' in session and session['authenticated'] == True

def require_auth(f):
    def decorated_function(*args, **kwargs):
        if not is_authenticated():
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Before request handler for session cleanup
@app.before_request
def cleanup_expired_sessions_before_request():
    """Clean up expired sessions before each request (occasionally)"""
    # Only run cleanup occasionally to avoid performance impact
    if random.random() < 0.1:  # 10% chance
        check_and_cleanup_expired_sessions()

# Routes
@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/favicon.ico')
def favicon():
    """Favicon route to prevent 404 errors"""
    return '', 204

@app.route('/api/entries')
def get_entries():
    """Get production entries with optional filtering"""
    try:
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        application = request.args.get('application')
        quality_status = request.args.get('quality_status')
        prb_only = request.args.get('prb_only', 'false').lower() == 'true'
        hiim_only = request.args.get('hiim_only', 'false').lower() == 'true'
        time_loss_only = request.args.get('time_loss_only', 'false').lower() == 'true'

        # Count how many row-level filters are active
        # All three filters (PRB, HIIM, Time Loss) are now treated as row-level filters
        # for individual row display when used alone
        active_row_filters = [prb_only, hiim_only, time_loss_only]
        active_row_filters_count = sum(1 for f in active_row_filters if f)

        # We use row-level (independent row) mode if EXACTLY one filter is selected
        use_row_level_filtering = (active_row_filters_count == 1)
        row_type_filter = None

        if use_row_level_filtering:
            if prb_only:
                row_type_filter = 'prb'
            elif hiim_only:
                row_type_filter = 'hiim'
            elif time_loss_only:
                row_type_filter = 'time_loss'

        # Retrieval strategy
        if use_row_level_filtering:
            # Single filter -> get individual rows filtered by the specific type, but enrich with complete data
            if application:
                # Get individual rows filtered by the specific row type (PRB, HIIM, or Time Loss)
                filtered_rows = entry_manager.get_individual_rows_by_application(application, start_date, end_date, row_type_filter)
                # Also get all rows for the same date/application to enrich the filtered rows
                all_rows = entry_manager.get_individual_rows_by_application(application, start_date, end_date)
            else:
                filtered_rows = entry_manager.get_all_individual_rows(row_type_filter)
                all_rows = entry_manager.get_all_individual_rows()
            
            # Group all rows by date and application for enrichment
            grouped_rows = {}
            for row in all_rows:
                key = f"{row.get('date')}_{row.get('application_name')}"
                if key not in grouped_rows:
                    grouped_rows[key] = []
                grouped_rows[key].append(row)
            
            # Enrich each filtered row with data from other rows in the same group
            enriched_entries = []
            for filtered_row in filtered_rows:
                key = f"{filtered_row.get('date')}_{filtered_row.get('application_name')}"
                enriched_row = filtered_row.copy()
                
                # Add data from other rows in the same group
                if key in grouped_rows:
                    for other_row in grouped_rows[key]:
                        # Add PRB data if not present
                        if not enriched_row.get('prb_id_number') and other_row.get('prb_id_number'):
                            enriched_row['prb_id_number'] = other_row.get('prb_id_number')
                            enriched_row['prb_id_status'] = other_row.get('prb_id_status', '')
                            enriched_row['prb_link'] = other_row.get('prb_link', '')
                        
                        # Add HIIM data if not present
                        if not enriched_row.get('hiim_id_number') and other_row.get('hiim_id_number'):
                            enriched_row['hiim_id_number'] = other_row.get('hiim_id_number')
                            enriched_row['hiim_id_status'] = other_row.get('hiim_id_status', '')
                            enriched_row['hiim_link'] = other_row.get('hiim_link', '')
                        
                        # Add Time Loss data if not present
                        if not enriched_row.get('time_loss') and other_row.get('time_loss'):
                            enriched_row['time_loss'] = other_row.get('time_loss')
                        
                        # Add Issue data if not present
                        if not enriched_row.get('issue_description') and other_row.get('issue_description'):
                            enriched_row['issue_description'] = other_row.get('issue_description')
                        
                        # Add other common fields if not present
                        for field in ['prc_mail_text', 'prc_mail_status', 'quality_status', 'remarks', 'day']:
                            if not enriched_row.get(field) and other_row.get(field):
                                enriched_row[field] = other_row.get(field)
                
                enriched_entries.append(enriched_row)
            
            all_entries = enriched_entries
        else:
            # Grouped mode (no row-level filters or multi-filter AND mode)
            if application:
                all_entries = entry_manager.get_entries_by_application(application, start_date, end_date)
            else:
                all_entries = entry_manager.get_all_entries()
        
        # Helper functions for filtering
        def has_prb(ent):
            if ent.get('prb_id_number'):
                return True
            prbs = ent.get('prbs') or []
            return any(p and (p.get('prb_id_number') or p.get('prb_id')) for p in prbs)

        def has_hiim(ent):
            if ent.get('hiim_id_number'):
                return True
            hiims = ent.get('hiims') or []
            return any(h and (h.get('hiim_id_number') or h.get('hiim_id')) for h in hiims)

        def has_time_loss(ent):
            # Check top-level time_loss field for meaningful values
            top_level_time_loss = ent.get('time_loss', '').strip()
            if top_level_time_loss and top_level_time_loss.upper() not in ['N/A', 'NA', 'NONE', 'NULL']:
                return True
            
            # Check issues array for meaningful time_loss values
            issues = ent.get('issues') or []
            return any(i and i.get('time_loss', '').strip() and 
                      i.get('time_loss', '').strip().upper() not in ['N/A', 'NA', 'NONE', 'NULL'] 
                      for i in issues)

        # Apply remaining filters (non-date, non-application filters)
        filtered_entries = []
        for entry in all_entries:
            # Date filters (only needed if not already filtered at database level)
            if not application and not use_row_level_filtering:  # Only apply date filters if we got all entries and not using row-level filtering
                if start_date:
                    entry_date = convert_date_string(entry.get('date', ''))
                    if entry_date < datetime.strptime(start_date, '%Y-%m-%d').date():
                        continue
                if end_date:
                    entry_date = convert_date_string(entry.get('date', ''))
                    if entry_date > datetime.strptime(end_date, '%Y-%m-%d').date():
                        continue
            
            # Application filter (only needed if we got all entries)
            if not application and 'application' in request.args:
                requested_app = request.args.get('application')
                if requested_app.lower() not in entry.get('application_name', '').lower():
                    continue
            
            # Quality status filter
            if quality_status:
                if entry.get('quality_status') != quality_status:
                    continue
            
            # Row-level single-filter case: apply the specific filter to the complete entry
            if use_row_level_filtering:
                # Apply the specific single filter to the complete entry
                if prb_only and not has_prb(entry):
                    continue
                if hiim_only and not has_hiim(entry):
                    continue
                if time_loss_only and not has_time_loss(entry):
                    continue
                filtered_entries.append(entry)
                continue

            # Multi-filter AND logic (or zero filters -> no extra constraints)
            # Helper checks (already defined above)

            # Apply AND conditions only for the filters that are active
            # Note: time_loss_only is now handled at row level, so skip entry-level filtering for it
            if prb_only and not has_prb(entry):
                continue
            if hiim_only and not has_hiim(entry):
                continue
            # time_loss_only filtering is now handled at database row level, not here

            filtered_entries.append(entry)
        
        # Sort by date descending, then by created_at descending
        filtered_entries.sort(key=lambda x: (
            convert_date_string(x.get('date', '1900-01-01')),
            x.get('created_at', '1900-01-01T00:00:00')
        ), reverse=True)
        
        # Create response with cache-busting headers to ensure fresh data
        response = jsonify(filtered_entries)
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entries/<int:entry_id>')
def get_entry(entry_id):
    """Get a specific production entry by ID"""
    try:
        application = request.args.get('application', '').upper()
        logger.info(f"API GET /api/entries/{entry_id} called with application={application}")
        entry = entry_manager.get_entry_by_id(entry_id, application if application else None)
        if entry:
            logger.info(f"Entry found: id={entry_id}, application={application}")
            response = jsonify(entry)
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response
        else:
            logger.warning(f"Entry NOT found: id={entry_id}, application={application}")
            return jsonify({'error': 'Entry not found'}), 404
    except Exception as e:
        logger.error(f"Error fetching entry {entry_id} for application {application}: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/entries', methods=['POST'])
@require_auth
def create_entry():
    """Create a new production entry with independent rows"""
    try:
        data = request.get_json()
        
        # Validate basic entry data
        is_valid, error_msg = validate_entry_data(data)
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        # Validate independent row constraints
        is_valid, error_msg = validate_independent_row_constraints(data)
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        # Serialize duplicate date/application check & create to avoid race
        with _create_entry_lock:
            application_name = data['application_name']
            existing_entries = entry_manager.get_entries_by_application(application_name)
            for existing_entry in existing_entries:
                if existing_entry.get('date') == data['date']:
                    return jsonify({'error': f'An entry already exists for {application_name} on {data["date"]}'}), 400
            # Create new entry
            entry = entry_manager.create_entry(data)
        
        if entry:
            return jsonify(entry), 201
        else:
            return jsonify({'error': 'Failed to create entry'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entries/<int:entry_id>', methods=['PUT'])
@require_auth
def update_entry(entry_id):
    """Update an existing production entry"""
    try:
        data = request.get_json()
        logger.info(f"API update_entry called for id={entry_id}, data={data}")

        # Get existing entry - search all databases
        existing_entry = entry_manager.get_entry_by_id(entry_id)
        if not existing_entry:
            logger.warning(f"Entry not found for id={entry_id}")
            return jsonify({'error': 'Entry not found'}), 404

        existing_application = existing_entry.get('application_name')

        # Check for duplicate entry if date or application is being changed
        if 'date' in data or 'application_name' in data:
            new_date = data.get('date', existing_entry.get('date'))
            new_application = data.get('application_name', existing_application)

            # Check if another entry exists for this date and application (excluding current entry)
            app_entries = entry_manager.get_entries_by_application(new_application)
            for entry in app_entries:
                if (entry.get('id') != entry_id and entry.get('date') == new_date):
                    logger.warning(f"Duplicate entry found for application={new_application} on date={new_date}")
                    return jsonify({'error': f'An entry already exists for {new_application} on {new_date}'}), 400

        # Validate entry data using the updated validation function
        merged_data = dict(existing_entry)
        merged_data.update(data)
        is_valid, error_message = validate_entry_data(merged_data)
        if not is_valid:
            logger.warning(f"Validation failed for id={entry_id}: {error_message}")
            return jsonify({'error': error_message}), 400

        # Update entry - pass the application name for database targeting
        updated_entry = entry_manager.update_entry(entry_id, data, existing_application)
        logger.info(f"Update result for id={entry_id}: {updated_entry}")

        # For XVA, return fresh entries list for immediate UI update
        if existing_application == 'XVA':
            fresh_entries = entry_manager.get_entries_by_application('XVA')
            response = jsonify({'updated_entry': updated_entry, 'fresh_entries': fresh_entries})
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response

        if updated_entry:
            return jsonify(updated_entry)
        else:
            logger.error(f"Failed to update entry for id={entry_id}")
            return jsonify({'error': 'Failed to update entry'}), 500
    except Exception as e:
        logger.error(f'Exception in API update_entry id={entry_id}: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/entries/<int:entry_id>', methods=['DELETE'])
@require_auth
def delete_entry(entry_id):
    """Delete a production entry"""
    try:
        # Get application from query param for efficient lookup
        application = request.args.get('application')
        # Delete entry from specific application database if provided
        success = entry_manager.delete_entry(entry_id, application)
        if application == 'XVA':
            fresh_entries = entry_manager.get_entries_by_application('XVA')
            response = jsonify({'message': 'Entry deleted successfully', 'fresh_entries': fresh_entries})
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response
        if success:
            response = jsonify({'message': 'Entry deleted successfully'})
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response
        else:
            return jsonify({'error': 'Entry not found or failed to delete'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/excel')
@require_auth
def download_excel():
    """Download current application dashboard data as Excel file with exact table structure"""
    try:
        # Get current application from request parameters
        application = request.args.get('application', 'CVAR ALL')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Get entries for the selected application with any applied filters
        if application == 'All Applications':
            entries = entry_manager.get_all_entries()
        else:
            entries = entry_manager.get_entries_by_application(application, start_date, end_date)
        
        if not entries:
            return jsonify({'error': f'No data available for {application}'}), 404
        
        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{application} Dashboard Data"
        
        # Define column structure based on application (matching exactly what dashboard shows)
        if application in ['CVAR ALL', 'CVAR NYQ']:
            headers = ['Date', 'Day', 'PRC Mail', 'CP Alerts', 'Quality', 'Punctuality Issue Description', 'Time Loss', 'PRB ID', 'HIIM ID', 'Remarks']
            field_mappings = [
                'date', 'day', 
                lambda e: f"{e.get('prc_mail_text', '')} {e.get('prc_mail_status', '')}".strip(),
                lambda e: f"{e.get('cp_alerts_text', '')} {e.get('cp_alerts_status', '')}".strip(),
                'quality_status', 'issue_description', 'time_loss', 'prb_id_number', 'hiim_id_number', 'remarks'
            ]
        elif application == 'XVA':
            headers = ['Date', 'Day', 'Acq', 'Valo', 'Sensi', 'CF RA', 'Quality Legacy', 'Quality Target', 'Root Cause Application', 'Root Cause Type', 'XVA Remarks', 'PRB ID', 'HIIM ID', 'Time Loss']
            field_mappings = [
                'date', 'day', 
                lambda e: f"{e.get('acq_text', '')} {e.get('acq_status', '')}".strip(),
                lambda e: f"{e.get('valo_text', '')} {e.get('valo_status', '')}".strip(),
                lambda e: f"{e.get('sensi_text', '')} {e.get('sensi_status', '')}".strip(),
                lambda e: f"{e.get('cf_ra_text', '')} {e.get('cf_ra_status', '')}".strip(),
                'quality_legacy', 'quality_target', 'root_cause_application', 'root_cause_type', 'xva_remarks', 'prb_id_number', 'hiim_id_number', 'time_loss'
            ]
        elif application == 'REG':
            headers = ['Date', 'Day', 'Closing', 'Iteration', 'Issue', 'Action Taken and Update', 'Status', 'PRB', 'HIIM', 'Backlog Item']
            field_mappings = [
                'date', 'day', 'closing', 'iteration', 'reg_issue', 'action_taken_and_update', 'reg_status', 'reg_prb', 'reg_hiim', 'backlog_item'
            ]
        elif application == 'OTHERS':
            headers = ['Date', 'Day', 'BUSINESS CHAIN', 'TIMINGS', 'PUNTUALITY ISSUE', 'QUALITY', 'QUALITY ISSUE', 'PRB', 'HIIM']
            field_mappings = [
                'date', 'day', 'business_chain', 'timings', 'puntuality_issue', 'quality', 'quality_issue', 'others_prb', 'others_hiim'
            ]
        else:
            # Default for unknown applications
            headers = ['Date', 'Day', 'PRC Mail', 'CP Alerts', 'Quality', 'Punctuality Issue Description', 'Time Loss', 'PRB ID', 'HIIM ID', 'Remarks']
            field_mappings = [
                'date', 'day', 
                lambda e: f"{e.get('prc_mail_text', '')} {e.get('prc_mail_status', '')}".strip(),
                lambda e: f"{e.get('cp_alerts_text', '')} {e.get('cp_alerts_status', '')}".strip(),
                'quality_status', 'issue_description', 'time_loss', 'prb_id_number', 'hiim_id_number', 'remarks'
            ]        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data rows using the field mappings for this application
        for row, entry in enumerate(entries, 2):
            for col, field_mapping in enumerate(field_mappings, 1):
                if callable(field_mapping):
                    # Field mapping is a function (for combined fields)
                    value = field_mapping(entry)
                else:
                    # Field mapping is a simple field name
                    value = entry.get(field_mapping, '')
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with current timestamp and application name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_app_name = application.replace(' ', '_').replace('/', '_')
        filename = f"ProdVision_{safe_app_name}_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel file: {str(e)}'}), 500

@app.route('/api/stats')
def get_stats():
    """Get aggregated statistics for charts"""
    try:
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        application = request.args.get('application')
        quality_status = request.args.get('quality_status')
        prb_only = request.args.get('prb_only')
        hiim_only = request.args.get('hiim_only')
        time_loss_only = request.args.get('time_loss_only')
        
        # Get monthly and yearly filters
        years = request.args.getlist('year')
        months = request.args.getlist('month')
        
        # Get entries from appropriate database(s)
        if application:
            all_entries = entry_manager.get_entries_by_application(application, start_date, end_date)
        else:
            all_entries = entry_manager.get_all_entries()
        
        # Apply additional filters (date filters already applied when application is specified)
        entries = []
        for entry in all_entries:
            # Date range filters (only apply when getting all entries, not when application-specific)
            if not application:
                if start_date:
                    entry_date = convert_date_string(entry.get('date', ''))
                    if entry_date < datetime.strptime(start_date, '%Y-%m-%d').date():
                        continue
                if end_date:
                    entry_date = convert_date_string(entry.get('date', ''))
                    if entry_date > datetime.strptime(end_date, '%Y-%m-%d').date():
                        continue
            
            # Monthly and yearly filters
            if years or months:
                entry_date = convert_date_string(entry.get('date', ''))
                entry_year = str(entry_date.year)
                entry_month = str(entry_date.month)
                
                if years and entry_year not in years:
                    continue
                if months and entry_month not in months:
                    continue
            
            # Other filters
            if quality_status and entry.get('quality_status') != quality_status:
                continue
            if prb_only == 'true' and not entry.get('prb_id_number'):
                continue
            if hiim_only == 'true' and not entry.get('hiim_id_number'):
                continue
            if time_loss_only == 'true' and not entry.get('time_loss'):
                continue
            
            entries.append(entry)
        
        # Calculate statistics
        total_entries = len(entries)
        quality_counts = {'Red': 0, 'Yellow': 0, 'Green': 0}
        punctuality_counts = {'Red': 0, 'Yellow': 0, 'Green': 0}
        prb_counts = {'active': 0, 'closed': 0}
        hiim_counts = {'active': 0, 'closed': 0}
        app_counts = {}
        
        # Monthly breakdown for comparison charts
        monthly_quality = {}
        monthly_punctuality = {}
        monthly_prb = {}
        monthly_hiim = {}
        
        # First, initialize all selected months with zero data
        if years or months:
            selected_months = set()
            
            # Generate all possible month-year combinations from selected years and months
            if years and months:
                # Both years and months are selected - use combinations
                for year in years:
                    for month in months:
                        month_key = f"{year}-{int(month):02d}"
                        month_name = datetime(int(year), int(month), 1).strftime('%B %Y')
                        selected_months.add((month_key, month_name))
            elif years:
                # Only years selected - include all months for those years
                for year in years:
                    for month in range(1, 13):
                        month_key = f"{year}-{month:02d}"
                        month_name = datetime(int(year), month, 1).strftime('%B %Y')
                        selected_months.add((month_key, month_name))
            elif months:
                # Only months selected - include those months for all years in the data
                # Get year range from the entries
                if entries:
                    years_in_data = set(convert_date_string(entry.get('date', '1900-01-01')).year for entry in entries)
                    for year in years_in_data:
                        for month in months:
                            month_key = f"{year}-{int(month):02d}"
                            month_name = datetime(int(year), int(month), 1).strftime('%B %Y')
                            selected_months.add((month_key, month_name))
            
            # Initialize all selected months with zero data
            for month_key, month_name in selected_months:
                monthly_quality[month_key] = {
                    'month_name': month_name,
                    'Red': 0, 'Yellow': 0, 'Green': 0
                }
                monthly_punctuality[month_key] = {
                    'month_name': month_name,
                    'Red': 0, 'Yellow': 0, 'Green': 0
                }
                monthly_prb[month_key] = {
                    'month_name': month_name,
                    'active': 0, 'closed': 0
                }
                monthly_hiim[month_key] = {
                    'month_name': month_name,
                    'active': 0, 'closed': 0
                }
        
        for entry in entries:
            # Get month-year key for grouping
            entry_date = convert_date_string(entry.get('date', '1900-01-01'))
            month_key = f"{entry_date.year}-{entry_date.month:02d}"
            month_name = entry_date.strftime('%B %Y')
            
            # Initialize monthly data if not exists (for cases where no year/month filters are applied)
            if month_key not in monthly_quality:
                monthly_quality[month_key] = {
                    'month_name': month_name,
                    'Red': 0, 'Yellow': 0, 'Green': 0
                }
            if month_key not in monthly_punctuality:
                monthly_punctuality[month_key] = {
                    'month_name': month_name,
                    'Red': 0, 'Yellow': 0, 'Green': 0
                }
            if month_key not in monthly_prb:
                monthly_prb[month_key] = {
                    'month_name': month_name,
                    'active': 0, 'closed': 0
                }
            if month_key not in monthly_hiim:
                monthly_hiim[month_key] = {
                    'month_name': month_name,
                    'active': 0, 'closed': 0
                }
            
            # Quality counts
            quality_status = entry.get('quality_status')
            if quality_status:
                quality_counts[quality_status] += 1
                monthly_quality[month_key][quality_status] += 1
            
            # Calculate punctuality - handle XVA vs other applications differently
            application_name = entry.get('application_name', '').upper()
            if application_name == 'XVA':
                # For XVA: Punctuality is red if any of valo, sensi, or cf_ra is red
                valo_status = entry.get('valo_status')
                sensi_status = entry.get('sensi_status')
                cf_ra_status = entry.get('cf_ra_status')
                
                if valo_status == 'Red' or sensi_status == 'Red' or cf_ra_status == 'Red':
                    punctuality_counts['Red'] += 1
                    monthly_punctuality[month_key]['Red'] += 1
                elif valo_status == 'Yellow' or sensi_status == 'Yellow' or cf_ra_status == 'Yellow':
                    punctuality_counts['Yellow'] += 1
                    monthly_punctuality[month_key]['Yellow'] += 1
                elif valo_status == 'Green' or sensi_status == 'Green' or cf_ra_status == 'Green':
                    punctuality_counts['Green'] += 1
                    monthly_punctuality[month_key]['Green'] += 1
                else:
                    # If no XVA statuses are set, count as Yellow
                    punctuality_counts['Yellow'] += 1
                    monthly_punctuality[month_key]['Yellow'] += 1
            else:
                # For other applications: Use PRC Mail status
                prc_mail_status = entry.get('prc_mail_status')
                if prc_mail_status:
                    # Map old status values to new color scheme
                    if prc_mail_status in ['Red', 'red']:
                        punctuality_counts['Red'] += 1
                        monthly_punctuality[month_key]['Red'] += 1
                    elif prc_mail_status in ['Yellow', 'yellow', 'warning']:
                        punctuality_counts['Yellow'] += 1
                        monthly_punctuality[month_key]['Yellow'] += 1
                    elif prc_mail_status in ['Green', 'green', 'on-time']:
                        punctuality_counts['Green'] += 1
                        monthly_punctuality[month_key]['Green'] += 1
                    elif prc_mail_status in ['late']:
                        punctuality_counts['Red'] += 1
                        monthly_punctuality[month_key]['Red'] += 1
                    else:
                        # For any other status, count as Yellow
                        punctuality_counts['Yellow'] += 1
                        monthly_punctuality[month_key]['Yellow'] += 1
            
            # Count PRB statuses (legacy single + array rows)
            prb_id_status = entry.get('prb_id_status')
            if prb_id_status:
                if prb_id_status in prb_counts:
                    prb_counts[prb_id_status] += 1
                    monthly_prb[month_key][prb_id_status] += 1
            # Also inspect prbs array if present
            if isinstance(entry.get('prbs'), list):
                for prb in entry['prbs']:
                    if prb and prb.get('prb_id_status') in prb_counts:
                        prb_counts[prb['prb_id_status']] += 1
                        monthly_prb[month_key][prb['prb_id_status']] += 1

            # Count HIIM statuses (legacy single + array rows)
            hiim_id_status = entry.get('hiim_id_status')
            if hiim_id_status:
                if hiim_id_status in hiim_counts:
                    hiim_counts[hiim_id_status] += 1
                    monthly_hiim[month_key][hiim_id_status] += 1
            if isinstance(entry.get('hiims'), list):
                for hiim in entry['hiims']:
                    if hiim and hiim.get('hiim_id_status') in hiim_counts:
                        hiim_counts[hiim['hiim_id_status']] += 1
                        monthly_hiim[month_key][hiim['hiim_id_status']] += 1
            
            application_name = entry.get('application_name', 'Unknown')
            app_counts[application_name] = app_counts.get(application_name, 0) + 1
        
        # Convert monthly data to sorted list
        monthly_quality_list = sorted(monthly_quality.items(), key=lambda x: x[0])
        monthly_punctuality_list = sorted(monthly_punctuality.items(), key=lambda x: x[0])
        monthly_prb_list = sorted(monthly_prb.items(), key=lambda x: x[0])
        monthly_hiim_list = sorted(monthly_hiim.items(), key=lambda x: x[0])
        
        return jsonify({
            'total_entries': total_entries,
            'quality_distribution': quality_counts,
            'punctuality_distribution': punctuality_counts,
            'prb_distribution': prb_counts,
            'hiim_distribution': hiim_counts,
            'application_distribution': app_counts,
            'monthly_quality': monthly_quality_list,
            'monthly_punctuality': monthly_punctuality_list,
            'monthly_prb': monthly_prb_list,
            'monthly_hiim': monthly_hiim_list
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Authenticate user"""
    try:
        data = request.get_json()
        password = data.get('password', '')
        
        # Get stored password hash from settings
        stored_hash = entry_manager.get_setting('admin_password')
        if not stored_hash:
            return jsonify({'error': 'Authentication not configured'}), 500
        
        # Check password
        if bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8')):
            session['authenticated'] = True
            session.permanent = True
            return jsonify({'message': 'Authentication successful'})
        else:
            return jsonify({'error': 'Invalid password'}), 401
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """Logout user"""
    # Delete session file immediately on logout
    delete_current_session_file()
    
    # Clear session data
    session.pop('authenticated', None)
    session.clear()
    
    return jsonify({'message': 'Logged out successfully'})

@app.route('/api/auth/status')
def auth_status():
    """Check authentication status"""
    return jsonify({'authenticated': is_authenticated()})

@app.route('/api/admin/cleanup-sessions', methods=['POST'])
@require_auth
def manual_cleanup_sessions():
    """Manually clean up expired session files"""
    try:
        deleted_count = check_and_cleanup_expired_sessions()
        return jsonify({
            'message': f'Successfully cleaned up {deleted_count} expired session files'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/session-stats')
@require_auth
def session_file_stats():
    """Get session file statistics"""
    try:
        stats = get_session_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/xva/stats')
def get_xva_stats():
    """Get XVA-specific statistics for charts and tables"""
    try:
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        years = request.args.getlist('year')
        months = request.args.getlist('month')
        
        # Get all entries from SharePoint SQLite database and filter for XVA only
        all_entries = entry_manager.get_all_entries()
        
        # Apply filters for XVA entries only
        entries = []
        for entry in all_entries:
            # Only XVA entries
            if entry.get('application_name') != 'XVA':
                continue
            
            # Date range filters
            if start_date:
                entry_date = convert_date_string(entry.get('date', ''))
                if entry_date < datetime.strptime(start_date, '%Y-%m-%d').date():
                    continue
            if end_date:
                entry_date = convert_date_string(entry.get('date', ''))
                if entry_date > datetime.strptime(end_date, '%Y-%m-%d').date():
                    continue
            
            # Monthly and yearly filters
            if years or months:
                entry_date = convert_date_string(entry.get('date', ''))
                entry_year = str(entry_date.year)
                entry_month = str(entry_date.month)
                
                if years and entry_year not in years:
                    continue
                if months and entry_month not in months:
                    continue
            
            entries.append(entry)
        
        # Calculate XVA-specific statistics
        monthly_red_counts = {}
        root_cause_analysis = {}
        
        # Initialize monthly data for all selected months
        if years or months:
            from datetime import datetime
            selected_months = set()
            
            if years and months:
                for year in years:
                    for month in months:
                        month_key = f"{year}-{int(month):02d}"
                        month_name = datetime(int(year), int(month), 1).strftime('%B %Y')
                        selected_months.add((month_key, month_name))
            elif years:
                for year in years:
                    for month in range(1, 13):
                        month_key = f"{year}-{month:02d}"
                        month_name = datetime(int(year), month, 1).strftime('%B %Y')
                        selected_months.add((month_key, month_name))
            elif months:
                if entries:
                    years_in_data = set(entry.date.year for entry in entries)
                    for year in years_in_data:
                        for month in months:
                            month_key = f"{year}-{int(month):02d}"
                            month_name = datetime(int(year), int(month), 1).strftime('%B %Y')
                            selected_months.add((month_key, month_name))
            
            # Initialize all selected months with zero data
            for month_key, month_name in selected_months:
                monthly_red_counts[month_key] = {
                    'month_name': month_name,
                    'valo_red': 0,
                    'sensi_red': 0,
                    'cf_ra_red': 0,
                    'total_red': 0
                }
        
        for entry in entries:
            # Get month-year key for grouping
            entry_date = convert_date_string(entry.get('date', '1900-01-01'))
            month_key = f"{entry_date.year}-{entry_date.month:02d}"
            month_name = entry_date.strftime('%B %Y')
            
            # Initialize monthly data if not exists
            if month_key not in monthly_red_counts:
                monthly_red_counts[month_key] = {
                    'month_name': month_name,
                    'valo_red': 0,
                    'sensi_red': 0,
                    'cf_ra_red': 0,
                    'total_red': 0
                }
            
            # Check if entry is a red card (punctuality OR quality is red)
            is_red_card = False
            
            # Check punctuality statuses for red
            if (entry.get('valo_status') == 'Red' or 
                entry.get('sensi_status') == 'Red' or 
                entry.get('cf_ra_status') == 'Red'):
                is_red_card = True
            
            # Check quality status for red
            if (entry.get('quality_legacy') == 'Red' or 
                entry.get('quality_target') == 'Red'):
                is_red_card = True
            
            if is_red_card:
                # Count red occurrences by category
                if entry.get('valo_status') == 'Red':
                    monthly_red_counts[month_key]['valo_red'] += 1
                
                if entry.get('sensi_status') == 'Red':
                    monthly_red_counts[month_key]['sensi_red'] += 1
                
                if entry.get('cf_ra_status') == 'Red':
                    monthly_red_counts[month_key]['cf_ra_red'] += 1
                
                # Increment total red count for this month
                monthly_red_counts[month_key]['total_red'] += 1
            
            # Root cause analysis - count all entries with root cause data, not just red cards
            if entry.get('root_cause_application'):
                root_cause_app = entry.get('root_cause_application') or 'Unknown'
                root_cause_type = entry.get('root_cause_type') or 'Unknown'
                
                key = f"{root_cause_app}|{root_cause_type}"
                if key not in root_cause_analysis:
                    root_cause_analysis[key] = {
                        'root_cause_application': root_cause_app,
                        'root_cause_type': root_cause_type,
                        'count': 0
                    }
                root_cause_analysis[key]['count'] += 1
        
        # Convert monthly data to sorted list
        monthly_red_counts_list = sorted(monthly_red_counts.items(), key=lambda x: x[0])
        
        # Convert root cause analysis to list and calculate grand total
        root_cause_list = list(root_cause_analysis.values())
        grand_total = sum(item['count'] for item in root_cause_list)
        
        return jsonify({
            'monthly_red_counts': monthly_red_counts_list,
            'root_cause_analysis': root_cause_list,
            'grand_total': grand_total
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Initialize database and create default admin password
def initialize_database():
    """Initialize SharePoint SQLite database with default settings"""
    try:
        # Ensure database tables exist
        entry_manager._ensure_datasets_exist()
        
        # Create default admin password if not exists
        admin_password = entry_manager.get_setting('admin_password')
        if not admin_password:
            # Default password is 'admin123' - should be changed in production
            hashed_password = bcrypt.hashpw('admin123'.encode('utf-8'), bcrypt.gensalt())
            entry_manager.set_setting('admin_password', hashed_password.decode('utf-8'))
        
        # Clean up any existing expired session files on startup
        cleanup_expired_session_files()
        
    except Exception as e:
        logger.error("Database initialization error: %s", e, exc_info=True)
        raise

if __name__ == '__main__':
    with app.app_context():
        initialize_database()
    
    # Start session cleanup thread
    cleanup_thread = threading.Thread(target=periodic_session_cleanup, daemon=True)
    cleanup_thread.start()

    logger.info("Starting.....")
    logger.info("Access the dashboard at: http://%s:%s", HOST, PORT)
 #   logger.warning("Default admin password: admin123 (CHANGE THIS IN PRODUCTION)")
    logger.info("Session cleanup: Enabled (runs every 30 minutes)")
    logger.info("Press Ctrl+C to stop the server")
    app.run(debug=DEBUG, host=HOST, port=PORT)
